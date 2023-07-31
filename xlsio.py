
from glob import glob
from tqdm import tqdm

import datetime

import numpy as np
import pandas as pd
from scipy import ndimage, spatial, optimize, signal, stats

from skimage import io, morphology, exposure
from skvideo import io as vio
import cv2

import openpyxl

from joblib import Parallel, delayed

def get_xy_of_name(sheet,STR="計測日"):
    avg_coords = []
    for n,i in enumerate(sheet.rows):
        for m,j in enumerate(i):
            if j.value == STR:
                avg_coords.append([n,m])
    return avg_coords

def get_xy_of_dates(sheet,y_coord,x_ind=1):
    dates = []
    for i in range(x_ind,sheet.max_column):
        val = sheet.cell(row=y_coord+1, column=x_ind+i).value

        dates.append(val)
    xd    = range(len(dates))
    dates = np.vstack([xd,dates])
    dates = dates.T[np.argwhere(dates[1]!=None)][:,0]
    return dates

def get_data(sheet,y_coord):
    """retrieve image coordinates, tomato number and image name
       this is incomlete for multiple clusters"""

    dates = get_xy_of_dates(sheet,y_coord)
    xframe= [dates[0,0]+2,dates[-1,0]+3]
    ystart= y_coord+2

    fname = None

    yx0 = sheet.cell(row=ystart, column=dates[0,0]+1).value
    if yx0 == None:

        fname = []
        for i in range(xframe[0],xframe[1]):
            val = sheet.cell(row=ystart, column=i).value
            fname.append(val)
        fname = np.asarray(fname)
        ystart+=1

    n   = ystart
    val = 1
    num = []
    while val != None:
        val = sheet.cell(row=n, column=dates[0,0]+1).value
        n   +=1
        if val != None:
            num.append(val)

    yframe = [ystart,n-1]
    ROW = []
    for j in range(yframe[0],yframe[1]):
        COL = []
        for i in range(xframe[0],xframe[1]):
            val = sheet.cell(row=j, column=i).value
            if val == None:
                val=-1
            COL.append(val)

        ROW.append(COL)
    return np.asarray(num),dates[:,1],np.asarray(ROW),fname

def format_pos(arr):
    narr = []
    for i in arr:
        temp = []
        for j in i:
            a = j
            if "," in a:
                a = a.split(",")
                b = [int(a[0]),int(a[1])]
            elif "." in a:
                a = a.split(".")
                b = [int(a[0]),int(a[1])]
            else:
                b = [int(a),int(a)]
            temp.append(b)
        narr.append(temp)
    return np.asarray(narr)

def get_all_data(sheet):
    n_list = get_xy_of_name(sheet)

    _,_,color   ,_  = get_data(sheet,n_list[0][0])
    _,_,diameter,_  = get_data(sheet,n_list[1][0])
    tomato,DATE,position,fname = get_data(sheet,n_list[2][0])

    return [fname,tomato,DATE,format_pos(position),diameter,color]

def get_all_sheet_data(sheets):
    DATAS = []
    for i in tqdm(sheets.sheetnames):
        sheet    = sheets.get_sheet_by_name(i)
        DATAS.append(get_all_data(sheet))
    return DATAS

def color_date_pair(COLOR,DATE):
    RES = []
    for i in COLOR:
        mask = np.where(i!=-1)
        DAYS = convert_to_day(DATE[mask])
        RES.append(np.asarray([DAYS,i[mask]]))
    return RES
